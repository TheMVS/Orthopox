import os
import pandas as pd
from openpyxl import load_workbook
from sklearn.model_selection import RandomizedSearchCV, train_test_split
from sklearn.metrics import make_scorer, cohen_kappa_score, f1_score
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from scipy.stats import randint, uniform, expon

from SklearnLoader import SklearnLoader


class SklearnModelRandomSearch:
    def __init__(self, X, Y, n_iter=2, cv=10, test_size=0.1, random_state=42):
        self.X = X
        self.Y = Y
        self.n_iter = n_iter
        self.cv = cv
        self.test_size = test_size
        self.random_state = random_state

        # Define statistical distributions for model hyperparameters
        self.models_params = {
            'RandomForest': {
                'clf__n_estimators': randint(50, 500),  # Más rango para n_estimators
                'clf__max_depth': randint(5, 50),  # Más rango para max_depth
                'clf__min_samples_split': randint(2, 20),  # Mayor rango de min_samples_split
                'clf__min_samples_leaf': randint(1, 10),  # Mayor rango de min_samples_leaf
                'clf__max_features': ['auto', 'sqrt', 'log2', None],  # Añadido max_features
                'clf__bootstrap': [True, False],  # Opciones para bootstrap
                'clf__class_weight': [None, 'balanced', 'balanced_subsample']  # Añadido 'balanced_subsample'
            },
            'SVM': {
                'clf__C': expon(scale=100),  # Mantiene la distribución exponencial para C
                'clf__kernel': ['linear', 'rbf', 'poly', 'sigmoid'],  # Añadido kernel sigmoid
                'clf__degree': randint(2, 5),  # Para kernel 'poly', variamos el grado
                'clf__gamma': ['scale', 'auto'],  # Opciones de gamma
                'clf__class_weight': [None, 'balanced']  # Mantiene las opciones de class_weight
            },
            'LogisticRegression': {
                'clf__C': uniform(0.001, 500),  # Mayor rango para C
                'clf__solver': ['liblinear', 'lbfgs', 'saga', 'newton-cg', 'newton-cholesky'],
                # Corregido para evitar duplicados
                'clf__penalty': ['l1', 'l2', 'elasticnet', None],  # Añadido elasticnet y None
                'clf__class_weight': [None, 'balanced'],  # Mantiene las opciones de class_weight
                'clf__max_iter': randint(100, 500)  # Añadido para iteraciones máximas
            },
            'DecisionTree': {
                'clf__criterion': ['gini', 'entropy', 'log_loss'],  # Variación en el criterio
                'clf__splitter': ['best', 'random'],  # Añadido splitter
                'clf__max_depth': randint(5, 50),  # Mayor rango para max_depth
                'clf__min_samples_split': randint(2, 20),  # Mayor rango para min_samples_split
                'clf__min_samples_leaf': randint(1, 10),  # Mayor rango para min_samples_leaf
                'clf__class_weight': [None, 'balanced']  # Mantiene las opciones de class_weight
            },
            'KNN':
                {
                    'clf__n_neighbors': [1],  # Caso fijo para k=1
                    'clf__weights': ['uniform', 'distance'],  # Mantiene las opciones de weights
                    'clf__algorithm': ['auto', 'ball_tree', 'kd_tree', 'brute'],  # Mantiene las opciones de algorithm
                    'clf__leaf_size': randint(10, 50),  # Añadido leaf_size
                    'clf__p': [1, 2]  # Distancia de Minkowski, p=1 (Manhattan) o p=2 (Euclidea)
                },
            'KNN':
                {
                    'clf__n_neighbors': randint(2, 50),  # Rango aleatorio para n_neighbors
                    'clf__weights': ['uniform', 'distance'],  # Mantiene las opciones de weights
                    'clf__algorithm': ['auto', 'ball_tree', 'kd_tree', 'brute'],  # Mantiene las opciones de algorithm
                    'clf__leaf_size': randint(10, 50),  # Añadido leaf_size
                    'clf__p': [1, 2]  # Distancia de Minkowski, p=1 (Manhattan) o p=2 (Euclidea)
                },
            'AdaBoost': {
                'clf__n_estimators': randint(50, 500),  # Variación en el número de estimadores
                'clf__learning_rate': uniform(0.01, 1),  # Variación en learning rate
                'clf__algorithm': ['SAMME', 'SAMME.R']  # Opciones para algoritmo
            },
            'ExtraTrees': {
                'clf__n_estimators': randint(50, 500),  # Variación en n_estimators
                'clf__max_depth': randint(5, 50),  # Variación en max_depth
                'clf__min_samples_split': randint(2, 20),  # Variación en min_samples_split
                'clf__min_samples_leaf': randint(1, 10),  # Variación en min_samples_leaf
                'clf__max_features': ['auto', 'sqrt', 'log2', None],  # Opciones para max_features
                'clf__bootstrap': [True, False],  # Variación en bootstrap
                'clf__class_weight': [None, 'balanced', 'balanced_subsample']  # Opciones para class_weight
            },
            'MLP': {
                'clf__hidden_layer_sizes': [(50,), (100,), (50, 50), (100, 100)],
                # Variaciones en el tamaño de las capas
                'clf__activation': ['identity', 'logistic', 'tanh', 'relu'],  # Opciones de activación
                'clf__solver': ['lbfgs', 'sgd', 'adam'],  # Opciones de solver
                'clf__alpha': uniform(0.0001, 0.1),  # Regularización L2
                'clf__learning_rate': ['constant', 'invscaling', 'adaptive'],  # Opciones para learning_rate
                'clf__max_iter': randint(200, 1000)  # Variación en el número máximo de iteraciones
            },
            'GaussianNB': {
                'clf__var_smoothing': uniform(1e-9, 1e-7)  # Suavizado de varianza
            },
            'GradientBoosting': {
                'clf__n_estimators': randint(10, 100),  # Reducción del rango de n_estimators para menos iteraciones
                'clf__learning_rate': uniform(0.05, 0.3),  # Rango más acotado para el learning rate
                'clf__max_depth': randint(2, 10),  # Reducción en el rango de max_depth para árboles más pequeños
                'clf__subsample': uniform(0.6, 1),  # Subsample para entrenar con menos muestras, acelerando el proceso
                'clf__min_samples_split': randint(2, 10),  # Rango menor para min_samples_split
                'clf__min_samples_leaf': randint(1, 5)  # Rango menor para min_samples_leaf
            }
        }
        self.results = []

    def _evaluate_model(self, model, X_train, X_test, y_train, y_test):
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        kappa = cohen_kappa_score(y_test, y_pred)
        f1 = f1_score(y_test, y_pred, average='weighted')
        return kappa, f1

    def _save_partial_results(self, results, filename, file_format):
        df = pd.DataFrame(results)

        if file_format == 'excel':
            # Check if the file exists and load previous data if it does
            if os.path.exists(filename):
                # Load the workbook to find the last row
                with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                    workbook = load_workbook(filename)
                    sheet = workbook.active
                    startrow = sheet.max_row
                    df.to_excel(writer, index=False, header=False, startrow=startrow)
            else:
                # If file doesn't exist, create it with headers
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, header=True)

        elif file_format == 'csv':
            # Check if the file exists, append data if it does
            if os.path.exists(filename):
                df.to_csv(filename, mode='a', header=False, index=False)
            else:
                df.to_csv(filename, mode='w', header=True, index=False)

    def perform_search(self, filename='random_search_results.xlsx', file_format='excel'):
        X_train, X_test, y_train, y_test = train_test_split(self.X, self.Y, test_size=self.test_size,
                                                            random_state=self.random_state)

        for model_name, param_distributions in self.models_params.items():
            loader = SklearnLoader(model_name)
            model = loader.get_model()

            pipeline = Pipeline([
                ('scaler', StandardScaler()),
                ('clf', model)
            ])

            search = RandomizedSearchCV(pipeline, param_distributions, n_iter=self.n_iter, cv=self.cv,
                                        scoring=make_scorer(f1_score, average='weighted'),
                                        random_state=self.random_state, n_jobs=-1, return_train_score=True, verbose=1)

            search.fit(X_train, y_train)

            model_results = []
            for rank in range(self.n_iter):
                best_model = search.cv_results_['params'][rank]
                mean_test_score = search.cv_results_['mean_test_score'][rank]
                std_test_score = search.cv_results_['std_test_score'][rank]

                kappa, f1 = self._evaluate_model(search.best_estimator_, X_train, X_test, y_train, y_test)

                model_results.append({
                    'model': model_name,
                    'rank': rank + 1,
                    'mean_test_score_cv': mean_test_score,
                    'std_test_score_cv': std_test_score,
                    'best_params': best_model,
                    'kappa': kappa,
                    'f1_score': f1
                })

            # Save results of the current model to the file in the specified format
            self._save_partial_results(model_results, filename, file_format)

    def run_search_and_save(self, filename='random_search_results', file_format='excel'):
        # Determine file extension based on the chosen format
        if 'xlsx' in file_format:
            filename += '.xlsx'
        elif 'csv' in file_format:
            filename += '.csv'

        self.perform_search(filename, file_format)
